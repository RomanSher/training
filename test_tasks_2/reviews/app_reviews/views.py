import csv
from .models import *
from .serializers import *
from openpyxl import Workbook
from rest_framework import generics
from django.http import HttpResponse, HttpResponseBadRequest
from rest_framework.permissions import IsAuthenticatedOrReadOnly



def export_country(request):
    format = request.GET['format']
    if format not in ['xlsx', 'csv']:
        return HttpResponseBadRequest
    queryset = Сountry.objects.all()
    if format == 'xlsx':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=country.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Сountry'
        columns = ['id', 'name']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for country in queryset:
            row_num += 1
            row = [country.pk, country.name]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=country.csv'
        opts = queryset.model._meta
        writer = csv.writer(response)
        field_names = [field.name for field in opts.fields]
        writer.writerow(field_names)
        for obj in queryset:
            writer.writerow([getattr(obj, field) for field in field_names])
        return response


def export_manufacturer(request):
    format = request.GET['format']
    if format not in ['xlsx', 'csv']:
        return HttpResponseBadRequest
    queryset = Manufacturer.objects.all()
    if format == 'xlsx':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=manufacturer.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Manufacturer'
        columns = ['id', 'name', 'country']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for manufacturer in queryset:
            row_num += 1
            row = [manufacturer.pk, manufacturer.name, manufacturer.country_id]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=manufacturer.csv'
        opts = queryset.model._meta
        writer = csv.writer(response)
        field_names = [field.name for field in opts.fields]
        writer.writerow(field_names)
        for obj in queryset:
            writer.writerow([getattr(obj, field) for field in field_names])
        return response


def export_car(request):
    format = request.GET['format']
    if format not in ['xlsx', 'csv']:
        return HttpResponseBadRequest
    queryset = Car.objects.all()
    if format == 'xlsx':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=car.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Car'
        columns = ['id', 'name', 'manufacturer', 'year_release', 'year_end']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for car in queryset:
            row_num += 1
            row = [car.pk, car.name, car.manufacturer_id, car.year_release, car.year_end]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=car.csv'
        opts = queryset.model._meta
        writer = csv.writer(response)
        field_names = [field.name for field in opts.fields]
        writer.writerow(field_names)
        for obj in queryset:
            writer.writerow([getattr(obj, field) for field in field_names])
        return response


def export_comment(request):
    format = request.GET['format']
    if format not in ['xlsx', 'csv']:
        return HttpResponseBadRequest
    queryset = Comment.objects.all()
    if format == 'xlsx':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=comment.xlsx'
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Comment'
        columns = ['id', 'email', 'car', 'created_at', 'comment']
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for comment in queryset:
            row_num += 1
            row = [comment.pk, comment.email, comment.car_id, comment.created_at, comment.comment]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=comment.csv'
        opts = queryset.model._meta
        writer = csv.writer(response)
        field_names = [field.name for field in opts.fields]
        writer.writerow(field_names)
        for obj in queryset:
            writer.writerow([getattr(obj, field) for field in field_names])
        return response


class СountryAPICreateRead(generics.ListCreateAPIView):
    queryset = Сountry.objects.all()
    serializer_class = СountrySerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)



class ManufacturerAPICreateRead(generics.ListCreateAPIView):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)


class CarAPICreateRead(generics.ListCreateAPIView):
    queryset = Car.objects.all()
    serializer_class = CarSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)


class CommentAPICreateRead(generics.ListCreateAPIView):
    queryset = Comment.objects.all()
    serializer_class = CommentSerializer


class СountryAPIUpdateDelete(generics.RetrieveUpdateDestroyAPIView):
    queryset = Сountry.objects.all()
    serializer_class = СountrySerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)


class ManufacturerAPIUpdateDelete(generics.RetrieveUpdateDestroyAPIView):
    queryset = Manufacturer.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)


class CarAPIUpdateDelete(generics.RetrieveUpdateDestroyAPIView):
    queryset = Car.objects.all()
    serializer_class = CarSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)


class CommentAPIUpdateDelete(generics.RetrieveUpdateDestroyAPIView):
    queryset = Comment.objects.all()
    serializer_class = CommentSerializer
    permission_classes = (IsAuthenticatedOrReadOnly,)
