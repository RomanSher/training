import csv
from django.http import HttpResponse
from .models import *
from import_export import resources
from openpyxl import Workbook


class DataMixin:

        def export_xlsx(self, query, model):
            queryset = model.objects.all()
            columns = [field.name for field in model._meta.get_fields()]
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename={query}.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = query
            row_num = 1

            for col_num, column_title in enumerate(columns[1:], 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title

            for obj in queryset:
                row_num += 1
                row = [getattr(obj, field) for field in columns[1:]]

                for col_num, cell_value in enumerate(row, 1):
                    if not isinstance(cell_value, str | datetime.date | int):
                        cell_value = cell_value.name
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value

            workbook.save(response)
            return response


        def export_csv(self, query, model):
            queryset = model.objects.all()
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename={query}.csv'
            opts = queryset.model._meta
            writer = csv.writer(response)
            field_names = [field.name for field in opts.fields]
            writer.writerow(field_names)
            for obj in queryset:
                writer.writerow([getattr(obj, field) for field in field_names])
            return response



# Закомментированные строки - это решение с помощью библиотеки django import-export

# class СountryResource(resources.ModelResource):
#     class Meta:
#         model = Сountry
#
# class ManufacturerResource(resources.ModelResource):
#     class Meta:
#         model = Manufacturer
#
# class СarResource(resources.ModelResource):
#     class Meta:
#         model = Car
#
# class CommentResource(resources.ModelResource):
#     class Meta:
#         model = Comment
#
#
# class DataMixin:
#
#     def export_xlsx(self, query, member_resource):
#         dataset = member_resource.export()
#         response = HttpResponse(dataset.xlsx, content_type='application/vnd.ms-excel')
#         response['Content-Disposition'] = f'attachment; filename="{query}.xlsx"'
#         return response
#
#
#     def export_csv(self, query, member_resource):
#         dataset = member_resource.export()
#         response = HttpResponse(dataset.csv, content_type='text/csv')
#         response['Content-Disposition'] = f'attachment; filename="{query}.csv"'
#         return response