from django.http import HttpResponse
from django.views.generic import ListView
from rest_framework import generics
from .models import News
from .serializers import NewsSerializer
from openpyxl import Workbook
from openpyxl.drawing.image import Image


class NewsListView(ListView):
    model = News
    template_name = 'news/news_list.html'


class NewsAPIView(generics.ListAPIView):
    queryset = News.objects.all()
    serializer_class = NewsSerializer


class NewsDetailAPIView(generics.ListAPIView):
    model = News
    serializer_class = NewsSerializer

    def get_queryset(self):
        news_id = self.kwargs['id']
        queryset = self.model.objects.filter(id=news_id)
        return queryset


def export(request):
    queryset = News.objects.all()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=news.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'News'
    columns = [
        'id',
        'title',
        'description',
        'created_at',
        'picture'
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for news in queryset:
        row_num += 1
        picture = Image(news.picture)
        picture.height = 20
        picture.width = 65
        row = [
            news.pk,
            news.title,
            news.description,
            news.created_at,
            picture
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            if isinstance(cell_value, Image):
                worksheet.add_image(cell_value, cell.coordinate)
            else:
                cell.value = cell_value

    workbook.save(response)
    return response